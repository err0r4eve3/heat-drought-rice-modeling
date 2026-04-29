"""Raw-data inventory scanning and metadata extraction."""

from __future__ import annotations

import csv
import json
import zipfile
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any


SUMMARY_COLUMNS = [
    "path",
    "file_name",
    "suffix",
    "size_mb",
    "modified_time",
    "guessed_category",
    "readable",
    "error_message",
]

SUPPORTED_SUFFIXES = {
    ".csv",
    ".xlsx",
    ".xls",
    ".nc",
    ".tif",
    ".tiff",
    ".shp",
    ".gpkg",
    ".geojson",
    ".parquet",
    ".hdf",
    ".h5",
    ".zip",
    ".json",
}

RAW_CATEGORIES = {
    "climate",
    "remote_sensing",
    "crop_mask",
    "phenology",
    "statistics",
    "boundary",
    "irrigation",
    "soil",
    "water",
    "references",
}


@dataclass(frozen=True)
class InventoryRecord:
    """Metadata for one file under data/raw."""

    path: str
    file_name: str
    suffix: str
    size_mb: float
    modified_time: str
    guessed_category: str
    readable: bool
    error_message: str
    details: dict[str, Any]

    def summary_row(self) -> dict[str, Any]:
        """Return the flat CSV row for this record."""

        return {column: getattr(self, column) for column in SUMMARY_COLUMNS}


@dataclass(frozen=True)
class InventoryResult:
    """Inventory scan result."""

    raw_dir: Path
    records: list[InventoryRecord]
    warnings: list[str]


def build_inventory(raw_dir: str | Path) -> InventoryResult:
    """Scan raw data files and extract lightweight metadata."""

    root = Path(raw_dir).expanduser().resolve()
    warnings: list[str] = []
    records: list[InventoryRecord] = []

    if not root.exists():
        warnings.append(f"Raw data directory does not exist: {root}")
        return InventoryResult(raw_dir=root, records=records, warnings=warnings)

    for file_path in sorted(path for path in root.rglob("*") if path.is_file()):
        record = _build_record(file_path, root)
        if not record.readable:
            warnings.append(f"{record.path}: {record.error_message}")
        records.append(record)

    return InventoryResult(raw_dir=root, records=records, warnings=warnings)


def write_inventory_outputs(
    inventory: InventoryResult,
    processed_dir: str | Path,
    reports_dir: str | Path,
) -> dict[str, Path]:
    """Write inventory CSV, detailed JSON, and Markdown report."""

    processed = Path(processed_dir).expanduser().resolve()
    reports = Path(reports_dir).expanduser().resolve()
    processed.mkdir(parents=True, exist_ok=True)
    reports.mkdir(parents=True, exist_ok=True)

    csv_path = processed / "data_inventory.csv"
    json_path = processed / "data_inventory_detailed.json"
    report_path = reports / "data_inventory.md"

    _write_summary_csv(inventory.records, csv_path)
    _write_detailed_json(inventory, json_path)
    report_path.write_text(render_inventory_report(inventory), encoding="utf-8")

    return {"csv": csv_path, "json": json_path, "report": report_path}


def render_inventory_report(inventory: InventoryResult) -> str:
    """Render a concise Markdown inventory report."""

    lines = [
        "# Data Inventory",
        "",
        f"- Generated at: {datetime.now().isoformat(timespec='seconds')}",
        f"- Raw data directory: `{inventory.raw_dir}`",
        f"- Total files: {len(inventory.records)}",
        f"- Readable files: {sum(record.readable for record in inventory.records)}",
        f"- Warning count: {len(inventory.warnings)}",
        "",
    ]

    if not inventory.records:
        lines.extend(["No raw data files found.", ""])
    else:
        lines.extend(
            [
                "| category | suffix | files | readable |",
                "| --- | --- | ---: | ---: |",
            ]
        )
        for (category, suffix), group in _group_records(inventory.records).items():
            readable_count = sum(record.readable for record in group)
            lines.append(f"| {category} | {suffix} | {len(group)} | {readable_count} |")
        lines.append("")

    if inventory.warnings:
        lines.extend(["## Warnings", ""])
        lines.extend(f"- {warning}" for warning in inventory.warnings)
        lines.append("")

    return "\n".join(lines)


def _build_record(file_path: Path, raw_dir: Path) -> InventoryRecord:
    """Build one inventory record and capture read errors."""

    suffix = file_path.suffix.lower()
    stat = file_path.stat()
    guessed_category = _guess_category(file_path, raw_dir)
    details: dict[str, Any] = {"supported_type": suffix in SUPPORTED_SUFFIXES}
    readable = False
    error_message = ""

    try:
        details.update(_inspect_file(file_path, suffix))
        readable = True
    except Exception as exc:  # noqa: BLE001 - inventory must not abort on a bad file
        error_message = f"{type(exc).__name__}: {exc}"

    return InventoryRecord(
        path=str(file_path),
        file_name=file_path.name,
        suffix=suffix,
        size_mb=round(stat.st_size / (1024 * 1024), 4),
        modified_time=datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds"),
        guessed_category=guessed_category,
        readable=readable,
        error_message=error_message,
        details=details,
    )


def _inspect_file(file_path: Path, suffix: str) -> dict[str, Any]:
    """Dispatch metadata extraction by file suffix."""

    if suffix == ".csv":
        return _inspect_csv(file_path)
    if suffix in {".xlsx", ".xls"}:
        return _inspect_excel(file_path)
    if suffix == ".nc":
        return _inspect_netcdf(file_path)
    if suffix in {".tif", ".tiff"}:
        return _inspect_geotiff(file_path)
    if suffix in {".shp", ".gpkg", ".geojson"}:
        return _inspect_vector(file_path)
    if suffix == ".parquet":
        return _inspect_parquet(file_path)
    if suffix in {".hdf", ".h5"}:
        return _inspect_hdf(file_path)
    if suffix == ".zip":
        return _inspect_zip(file_path)
    if suffix == ".json":
        return _inspect_json(file_path)
    return {"note": "Unsupported file type; basic file metadata only."}


def _inspect_csv(file_path: Path) -> dict[str, Any]:
    """Read CSV headers, row count, and first rows using a streaming reader."""

    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "latin1"):
        try:
            with file_path.open("r", encoding=encoding, newline="") as file_obj:
                reader = csv.DictReader(file_obj)
                columns = reader.fieldnames or []
                sample: list[dict[str, Any]] = []
                row_count = 0
                for row in reader:
                    if row_count < 5:
                        sample.append(dict(row))
                    row_count += 1
            return {
                "encoding": encoding,
                "columns": columns,
                "row_count": row_count,
                "sample_rows": sample,
            }
        except UnicodeDecodeError as exc:
            last_error = exc
    if last_error is not None:
        raise last_error
    raise ValueError("CSV could not be read")


def _inspect_excel(file_path: Path) -> dict[str, Any]:
    """Read workbook metadata and first rows."""

    if file_path.suffix.lower() == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ImportError("openpyxl is required to inspect .xlsx files") from exc

        workbook = load_workbook(file_path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        header = [str(value) if value is not None else "" for value in next(rows, ())]
        sample = []
        for row_index, row in enumerate(rows):
            if row_index >= 5:
                break
            sample.append({header[i]: row[i] for i in range(min(len(header), len(row)))})
        row_count = max((sheet.max_row or 1) - 1, 0)
        workbook.close()
        return {
            "sheet_name": sheet.title,
            "columns": header,
            "row_count": row_count,
            "sample_rows": sample,
        }

    try:
        import pandas as pd
    except ImportError as exc:
        raise ImportError("pandas is required to inspect .xls files") from exc

    frame = pd.read_excel(file_path)
    return {
        "columns": [str(column) for column in frame.columns],
        "row_count": int(len(frame)),
        "sample_rows": frame.head(5).to_dict(orient="records"),
    }


def _inspect_netcdf(file_path: Path) -> dict[str, Any]:
    """Read NetCDF variables, dimensions, and basic coordinate ranges."""

    try:
        import xarray as xr
    except ImportError as exc:
        raise ImportError("xarray is required to inspect NetCDF files") from exc

    with xr.open_dataset(file_path, decode_times=True, mask_and_scale=False) as dataset:
        details: dict[str, Any] = {
            "variables": list(dataset.data_vars),
            "dimensions": {name: int(size) for name, size in dataset.sizes.items()},
        }
        for coord_name in ("time", "lat", "latitude", "y", "lon", "longitude", "x"):
            if coord_name in dataset.coords:
                coord = dataset.coords[coord_name]
                if coord.size:
                    values = coord.values
                    details[f"{coord_name}_range"] = [str(values.min()), str(values.max())]
        return details


def _inspect_geotiff(file_path: Path) -> dict[str, Any]:
    """Read GeoTIFF CRS, bounds, resolution, and band metadata."""

    try:
        import rasterio
    except ImportError as exc:
        raise ImportError("rasterio is required to inspect GeoTIFF files") from exc

    with rasterio.open(file_path) as dataset:
        return {
            "crs": str(dataset.crs) if dataset.crs else None,
            "bounds": list(dataset.bounds),
            "resolution": list(dataset.res),
            "band_count": int(dataset.count),
            "nodata": dataset.nodata,
        }


def _inspect_vector(file_path: Path) -> dict[str, Any]:
    """Read vector CRS, bounds, geometry type, feature count, and columns."""

    try:
        import geopandas as gpd
    except ImportError as exc:
        raise ImportError("geopandas is required to inspect vector files") from exc

    frame = gpd.read_file(file_path)
    geometry_types = sorted(str(value) for value in frame.geometry.geom_type.dropna().unique())
    return {
        "crs": str(frame.crs) if frame.crs else None,
        "bounds": list(frame.total_bounds) if len(frame) else None,
        "geometry_type": geometry_types,
        "feature_count": int(len(frame)),
        "columns": [str(column) for column in frame.columns],
    }


def _inspect_parquet(file_path: Path) -> dict[str, Any]:
    """Read Parquet schema and row count using pyarrow metadata."""

    try:
        import pyarrow.parquet as pq
    except ImportError as exc:
        raise ImportError("pyarrow is required to inspect Parquet files") from exc

    parquet_file = pq.ParquetFile(file_path)
    return {
        "columns": parquet_file.schema.names,
        "row_count": int(parquet_file.metadata.num_rows),
        "row_groups": int(parquet_file.metadata.num_row_groups),
    }


def _inspect_hdf(file_path: Path) -> dict[str, Any]:
    """Read top-level HDF5 groups and datasets."""

    try:
        import h5py
    except ImportError as exc:
        raise ImportError("h5py is required to inspect HDF files") from exc

    with h5py.File(file_path, "r") as hdf:
        return {"top_level_keys": sorted(str(key) for key in hdf.keys())}


def _inspect_zip(file_path: Path) -> dict[str, Any]:
    """Read ZIP member count and first file names."""

    with zipfile.ZipFile(file_path) as archive:
        names = archive.namelist()
    return {"file_count": len(names), "sample_files": names[:20]}


def _inspect_json(file_path: Path) -> dict[str, Any]:
    """Read JSON top-level type and keys where available."""

    with file_path.open("r", encoding="utf-8") as file_obj:
        payload = json.load(file_obj)
    details: dict[str, Any] = {"json_type": type(payload).__name__}
    if isinstance(payload, dict):
        details["top_level_keys"] = list(payload.keys())[:50]
    elif isinstance(payload, list):
        details["item_count"] = len(payload)
        details["sample_items"] = payload[:5]
    return details


def _guess_category(file_path: Path, raw_dir: Path) -> str:
    """Guess data category from the first directory under raw."""

    try:
        relative = file_path.relative_to(raw_dir)
    except ValueError:
        return "unknown"
    if relative.parts and relative.parts[0] in RAW_CATEGORIES:
        return relative.parts[0]
    return "unknown"


def _write_summary_csv(records: list[InventoryRecord], output_path: Path) -> None:
    """Write the flat inventory CSV."""

    with output_path.open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.DictWriter(file_obj, fieldnames=SUMMARY_COLUMNS)
        writer.writeheader()
        for record in records:
            writer.writerow(record.summary_row())


def _write_detailed_json(inventory: InventoryResult, output_path: Path) -> None:
    """Write detailed JSON inventory."""

    payload = {
        "summary": {
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "raw_dir": str(inventory.raw_dir),
            "total_files": len(inventory.records),
            "readable_files": sum(record.readable for record in inventory.records),
            "warning_count": len(inventory.warnings),
            "by_suffix": _count_by_suffix(inventory.records),
        },
        "warnings": inventory.warnings,
        "records": [asdict(record) for record in inventory.records],
    }
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _count_by_suffix(records: list[InventoryRecord]) -> dict[str, int]:
    """Count inventory records by suffix."""

    counts: dict[str, int] = {}
    for record in records:
        counts[record.suffix] = counts.get(record.suffix, 0) + 1
    return dict(sorted(counts.items()))


def _group_records(
    records: list[InventoryRecord],
) -> dict[tuple[str, str], list[InventoryRecord]]:
    """Group records by category and suffix for reporting."""

    grouped: dict[tuple[str, str], list[InventoryRecord]] = {}
    for record in records:
        key = (record.guessed_category, record.suffix or "(none)")
        grouped.setdefault(key, []).append(record)
    return dict(sorted(grouped.items()))
